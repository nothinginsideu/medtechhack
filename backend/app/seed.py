import json
import asyncio
from sqlalchemy.future import select
from sqlalchemy.future import select
from app.db.database import SessionLocal, engine
from app.models.partner import Partner
from app.models.service import Service
from app.models.price_document import PriceDocument
from app.models.price_item import PriceItem
from app.models.base import Base
import os

async def seed_data():
    # Для хакатона больше не пересоздаем схему автоматически, используем миграции
    pass

    async with SessionLocal() as session:
        # 1. Seed Partners
        # Для рабочего продукта убираем создание тестовых демо-клиник "Клиника А" и "Клиника Б"
        # partners_file = "/app/data/partners.json"
        # if os.path.exists(partners_file):
        #     with open(partners_file, "r", encoding="utf-8") as f:
        #         partners_data = json.load(f)
        #     
        #     for p_data in partners_data:
        #         stmt = select(Partner).where(Partner.name == p_data["name"])
        #         result = await session.execute(stmt)
        #         if not result.scalars().first():
        #             new_partner = Partner(
        #                 name=p_data["name"],
        #                 city=p_data["city"],
        #                 address=p_data["address"],
        #                 contact_phone=p_data.get("phone", None),
        #                 config=p_data.get("config", {})
        #             )
        #             session.add(new_partner)
        #     print("Partners seeded.")
        
        # 2. Seed Services from Справочник
        services_file = "/app/data/Справочник услуг.xlsx"
        import openpyxl
        if os.path.exists(services_file):
            wb = openpyxl.load_workbook(services_file, data_only=True)
            sheet = wb["Справочник"] if "Справочник" in wb.sheetnames else wb.active
            
            # Находим индексы нужных колонок в первой строке
            header = [cell.value for cell in sheet[1]]
            col_idx = {name: i for i, name in enumerate(header)}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name_ru = str(row[col_idx.get("Name_ru", 3)] or "").strip()
                if not name_ru or name_ru == "None":
                    continue
                
                specialty = str(row[col_idx.get("Специальность", 1)] or "").strip()
                code = str(row[col_idx.get("Code", 2)] or "").strip()
                tarificator_code = str(row[col_idx.get("TarificatrCode", 4)] or "").strip()
                
                # Проверим существует ли уже такая услуга
                stmt = select(Service).where(Service.name_ru == name_ru)
                result = await session.execute(stmt)
                if not result.scalars().first():
                    new_service = Service(
                        specialty=specialty if specialty != "None" else None,
                        code=code if code != "None" else None,
                        name_ru=name_ru,
                        tarificator_code=tarificator_code if tarificator_code != "None" else None
                    )
                    session.add(new_service)
            print("Services seeded.")

        await session.commit()
        print("Seeding completed successfully.")

if __name__ == "__main__":
    asyncio.run(seed_data())
